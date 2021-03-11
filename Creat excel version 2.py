# Batuhan Beel
# Create Excel
# 25.10.2020
from PyQt5 import QtCore, QtGui, QtWidgets
from openpyxl import load_workbook,Workbook
import os
from PyQt5.QtGui import *

from PyQt5 import QtCore, QtGui, QtWidgets
from PIL import Image,ImageFilter

class Ui_Form(object):
    def setupUi(self, Form):
        self.i = 0
        self.workbook = Workbook("Data.xlsx")
        self.worksheet = self.workbook.active
        if not os.path.isfile("Data.xlsx"):  # Checks if the Excel file exists
            self.workbook.save("Data.xlsx")  # Create Excel file if not exists
            wb = load_workbook(filename="Data.xlsx")
            ws = wb.active
            liste = ["Name", "Surname", "Gender", "Age", "City", "Telephone Number"]
            ws.append(liste)
            wb.save(filename="Data.xlsx")
        Form.setObjectName("Form")
        Form.resize(900, 570)
        Form.setMaximumSize(900,570)
        Form.setMinimumSize(900,570)
        Form.setStyleSheet("""
                    background:  #F28068;

                    """)
        self.lineEdit_name = QtWidgets.QLineEdit(Form)
        self.lineEdit_name.setGeometry(QtCore.QRect(30, 50, 361, 31))
        self.lineEdit_name.setObjectName("lineEdit_name")
        self.lineEdit_name.setStyleSheet("""
               background:white;
               """)
        self.label_name = QtWidgets.QLabel(Form)
        self.label_name.setGeometry(QtCore.QRect(23, 20, 61, 21))
        self.label_name.setLayoutDirection(QtCore.Qt.LeftToRight)
        self.label_name.setAutoFillBackground(False)
        self.label_name.setAlignment(QtCore.Qt.AlignCenter)
        self.label_name.setObjectName("label_name")
        self.label_name.setFont(QFont("Arial", 10, QtGui.QFont.Bold))
        self.label_surname = QtWidgets.QLabel(Form)
        self.label_surname.setGeometry(QtCore.QRect(30, 90, 61, 21))
        self.label_surname.setAlignment(QtCore.Qt.AlignCenter)
        self.label_surname.setObjectName("label_surname")
        self.label_surname.setFont(QFont("Arial", 10, QtGui.QFont.Bold))
        self.lineEdit_surname = QtWidgets.QLineEdit(Form)
        self.lineEdit_surname.setGeometry(QtCore.QRect(30, 120, 361, 31))
        self.lineEdit_surname.setObjectName("lineEdit_surname")
        self.lineEdit_surname.setStyleSheet("""
         background:white;
         """)
        self.label_gender = QtWidgets.QLabel(Form)
        self.label_gender.setGeometry(QtCore.QRect(25, 160, 61, 21))
        self.label_gender.setAlignment(QtCore.Qt.AlignCenter)
        self.label_gender.setObjectName("label_gender")
        self.label_gender.setFont(QFont("Arial", 10, QtGui.QFont.Bold))
        self.lineEdit_gender = QtWidgets.QLineEdit(Form)
        self.lineEdit_gender.setGeometry(QtCore.QRect(30, 190, 361, 31))
        self.lineEdit_gender.setObjectName("lineEdit_gender")
        self.lineEdit_gender.setStyleSheet("""
        background:white;
        """)
        self.labelage4 = QtWidgets.QLabel(Form)
        self.labelage4.setGeometry(QtCore.QRect(15, 230, 61, 21))
        self.labelage4.setAlignment(QtCore.Qt.AlignCenter)
        self.labelage4.setObjectName("labelage4")
        self.labelage4.setFont(QFont("Arial", 10, QtGui.QFont.Bold))
        self.lineEdit_Age = QtWidgets.QLineEdit(Form)
        self.lineEdit_Age.setGeometry(QtCore.QRect(30, 260, 361, 31))
        self.lineEdit_Age.setObjectName("lineEdit_Age")
        self.lineEdit_Age.setStyleSheet("""
        background:white;
        """)
        self.label_city = QtWidgets.QLabel(Form)
        self.label_city.setGeometry(QtCore.QRect(15, 300, 61, 21))
        self.label_city.setAlignment(QtCore.Qt.AlignCenter)
        self.label_city.setObjectName("label_city")
        self.label_city.setFont(QFont("Arial", 10, QtGui.QFont.Bold))
        self.lineEdit_city = QtWidgets.QLineEdit(Form)
        self.lineEdit_city.setGeometry(QtCore.QRect(30, 330, 361, 31))
        self.lineEdit_city.setObjectName("lineEdit_city")
        self.lineEdit_city.setStyleSheet("""
        background:white;
        """)
        self.label_telno = QtWidgets.QLabel(Form)
        self.label_telno.setGeometry(QtCore.QRect(25, 370, 131, 21))
        self.label_telno.setAlignment(QtCore.Qt.AlignCenter)
        self.label_telno.setObjectName("label_telno")
        self.label_telno.setFont(QFont("Arial", 10, QtGui.QFont.Bold))
        self.lineEdit_telno = QtWidgets.QLineEdit(Form)
        self.lineEdit_telno.setGeometry(QtCore.QRect(30, 400, 361, 31))
        self.lineEdit_telno.setObjectName("lineEdit_telno")
        self.lineEdit_telno.setStyleSheet("""
        background:white;
        """)
        self.pushButton_result_1 = QtWidgets.QPushButton(Form)
        self.pushButton_result_1.setGeometry(QtCore.QRect(310, 450, 75, 23))
        self.pushButton_result_1.setStyleSheet("""
        background:#EFD6D0;
        """)
        self.pushButton_result_1.setObjectName("pushButton_result_1")
        self.lineEdit_name_surname = QtWidgets.QLineEdit(Form)
        self.lineEdit_name_surname.setGeometry(QtCore.QRect(460, 50, 361, 31))
        self.lineEdit_name_surname.setObjectName("lineEdit_name_surname")
        self.lineEdit_name_surname.setStyleSheet("""
        background:white;
        """)
        self.pushButton_result_2 = QtWidgets.QPushButton(Form)
        self.pushButton_result_2.setGeometry(QtCore.QRect(750, 110, 75, 23))
        self.pushButton_result_2.setObjectName("pushButton_result_2")
        self.pushButton_result_2.setStyleSheet("""
        background:#EFD6D0;
        """)
        self.label_information = QtWidgets.QLabel(Form)
        self.label_information.setGeometry(QtCore.QRect(470, 170, 341, 141))
        self.label_information.setText("")
        self.label_information.setAlignment(QtCore.Qt.AlignLeading|QtCore.Qt.AlignLeft|QtCore.Qt.AlignTop)
        self.label_information.setObjectName("label_information")
        self.label_information.setFont(QFont("Times New Roman",14))
        self.label_result_2 = QtWidgets.QLabel(Form)
        self.label_result_2.setGeometry(QtCore.QRect(690, 140, 131, 21))
        self.label_result_2.setText("")
        self.label_result_2.setAlignment(QtCore.Qt.AlignCenter)
        self.label_result_2.setObjectName("label_result_2")
        self.label_result = QtWidgets.QLabel(Form)
        self.label_result.setGeometry(QtCore.QRect(200, 490, 201, 20))
        self.label_result.setText("")
        self.label_result.setAlignment(QtCore.Qt.AlignCenter)
        self.label_result.setObjectName("label_result")
        self.label_name_surname = QtWidgets.QLabel(Form)
        self.label_name_surname.setGeometry(QtCore.QRect(470, 20, 101, 21))
        self.label_name_surname.setObjectName("label_name_surname")
        self.label_name_surname.setFont(QFont("Arial", 10, QtGui.QFont.Bold))
        self.label_name_surname_delete_user = QtWidgets.QLabel(Form)
        self.label_name_surname_delete_user.setGeometry(QtCore.QRect(470, 365, 101, 31))
        self.label_name_surname_delete_user.setObjectName("label_name_surname_delete_user")
        self.label_name_surname_delete_user.setFont(QFont("Arial", 10, QtGui.QFont.Bold))
        self.label_delete_user = QtWidgets.QLabel(Form)
        self.label_delete_user.setGeometry(QtCore.QRect(466, 340, 81, 20))
        self.label_delete_user.setAlignment(QtCore.Qt.AlignCenter)
        self.label_delete_user.setObjectName("label_delete_user")
        self.label_delete_user.setFont(QFont("Arial", 10, QtGui.QFont.Bold))
        self.lineEdit_name_surname_delete_user = QtWidgets.QLineEdit(Form)
        self.lineEdit_name_surname_delete_user.setGeometry(QtCore.QRect(470, 400, 361, 31))
        self.lineEdit_name_surname_delete_user.setObjectName("lineEdit_name_surname_delete_user")
        self.lineEdit_name_surname_delete_user.setStyleSheet("""
        background:white;
        """)
        self.pushButton_result_3 = QtWidgets.QPushButton(Form)
        self.pushButton_result_3.setStyleSheet("""
                background:#EFD6D0;
                """)
        self.pushButton_result_3.setFont(QFont("Arial",9,QFont.Bold))
        self.pushButton_result_2.setFont(QFont("Arial", 9, QFont.Bold))
        self.pushButton_result_1.setFont(QFont("Arial", 9, QFont.Bold))
        self.pushButton_result_3.setGeometry(QtCore.QRect(750, 450, 75, 23))
        self.pushButton_result_3.setObjectName("pushButton_result_3")
        self.label_result_3 = QtWidgets.QLabel(Form)
        self.label_result_3.setGeometry(QtCore.QRect(690, 490, 131, 21))
        self.label_result_3.setText("")
        self.label_result_3.setAlignment(QtCore.Qt.AlignCenter)
        self.label_result_3.setObjectName("label_result_3")

        self.pushButton_result_1.clicked.connect(self.excel_append)
        self.pushButton_result_2.clicked.connect(self.excel_search)
        self.pushButton_result_3.clicked.connect(self.excel_delete)

        self.retranslateUi(Form)
        QtCore.QMetaObject.connectSlotsByName(Form)

    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "Excel"))
        self.label_name.setText(_translate("Form", "Name"))
        self.label_surname.setText(_translate("Form", "Surname"))
        self.label_gender.setText(_translate("Form", "Gender"))
        self.labelage4.setText(_translate("Form", "Age"))
        self.label_city.setText(_translate("Form", "City"))
        self.label_telno.setText(_translate("Form", "Telephone Number"))
        self.pushButton_result_1.setText(_translate("Form", "Add"))
        self.pushButton_result_2.setText(_translate("Form", "Search"))
        self.label_name_surname.setText(_translate("Form", "Name-Surname"))
        self.label_name_surname_delete_user.setText(_translate("Form", "Name-Surname"))
        self.label_delete_user.setText(_translate("Form", "Delete User"))
        self.pushButton_result_3.setText(_translate("Form", "Delete"))

    def excel_append(self,Form):
        wb = load_workbook(filename="Data.xlsx")
        ws = wb.active
        try:
            if len(self.lineEdit_name.text()) != 0 and len(self.lineEdit_surname.text()) != 0 and int(self.lineEdit_Age.text()) > 0 and len(self.lineEdit_city.text()) != 0 and len(self.lineEdit_gender.text()) != 0 and len(self.lineEdit_telno.text()) == 11:
                list=[self.lineEdit_name.text(),self.lineEdit_surname.text(),self.lineEdit_gender.text(),self.lineEdit_Age.text(),self.lineEdit_city.text(),self.lineEdit_telno.text()]
                ws.append(list) # To add informations in cells
                wb.save(filename="Data.xlsx") # To update Excel file
                self.lineEdit_name.setText("")
                self.lineEdit_surname.setText("")
                self.lineEdit_Age.setText("")
                self.lineEdit_gender.setText("")
                self.lineEdit_telno.setText("")
                self.lineEdit_city.setText("")
                self.label_result.setText("Saved in Excel")
            else:
                self.label_result.setText("Error--->Please Check Your Information") # Error Message
        except:
            self.label_result.setText("Error--->Please Check Your Information")     # Error Message


    def excel_search(self,Form):
        wb = load_workbook(filename="Data.xlsx")
        ws = wb.active
        flag=0
        for i in range(1,ws.max_row+1):
            name         = str(ws.cell(i,1).value)
            surname      = str( ws.cell(i, 2).value)
            name_surname = name+" "+surname
            if self.lineEdit_name_surname.text() == name_surname: # To search person's name in excel file
                self.label_information.setText("Name : {}\nSurname : {}\nGender : {}\nAge : {}\nCity : {}\nTelephone Number : {}".format(ws.cell(i,1).value,ws.cell(i,2).value,ws.cell(i,3).value,ws.cell(i,4).value,ws.cell(i,5).value,ws.cell(i,6).value))
                flag=1
        if flag == 0:
            self.label_result_2.setText("Not Found")
            self.label_information.setText("")
    def excel_delete(self,Form):
        wb = load_workbook(filename="Data.xlsx")
        ws = wb.active
        flag=0
        for i in range(1,ws.max_row+1):
            name         = str(ws.cell(i,1).value)
            surname      = str( ws.cell(i, 2).value)
            name_surname = name+" "+surname
            if self.lineEdit_name_surname_delete_user.text() == name_surname: # To search person's name in excel file
                ws.delete_rows(i,1)
                wb.save(filename="Data.xlsx") # To update Excel file
                self.label_result_3.setText("User successfully deleted.")
                flag=1
        if flag == 0:
            self.label_result_3.setText("Not Found.")



if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Form = QtWidgets.QWidget()
    ui = Ui_Form()
    ui.setupUi(Form)
    Form.show()
    sys.exit(app.exec_())
